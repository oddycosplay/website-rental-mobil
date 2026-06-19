import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Black Box Testing"

# --- Styles ---
def hdr_style(ws, row, col, val, bg="1F3864", fg="FFFFFF", bold=True, sz=11, wrap=True, halign="center", valign="center"):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Calibri", bold=bold, color=fg, size=sz)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    return c

def cell_style(ws, row, col, val, bold=False, halign="left", valign="center", wrap=True, bg=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Calibri", bold=bold, size=10)
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

thin = Side(border_style="thin", color="AAAAAA")
bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

def apply_border(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row+1):
        for c in range(start_col, end_col+1):
            ws.cell(r, c).border = bdr

# =============================================
# TITLE
# =============================================
ws.merge_cells("A1:G1")
t = ws.cell(1, 1, "DOKUMEN PENGUJIAN BLACK BOX TESTING")
t.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
t.fill = PatternFill("solid", fgColor="1F3864")
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:G2")
s = ws.cell(2, 1, "Sistem Informasi Rental Mobil – Siliwangi Rental Trans Nusa")
s.font = Font(name="Calibri", bold=False, size=11, color="FFFFFF")
s.fill = PatternFill("solid", fgColor="2E4D87")
s.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

ws.row_dimensions[3].height = 8

# =============================================
# HEADER TABLE
# =============================================
headers = ["No", "Skenario Uji", "Test Case", "Hasil yang Diharapkan", "Hasil Pengujian", "Status", "Kesimpulan"]
col_widths = [5, 28, 42, 42, 42, 12, 18]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    hdr_style(ws, 4, i, h, bg="2E75B6", sz=10)
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[4].height = 30
apply_border(ws, 4, 4, 1, 7)

# =============================================
# DATA
# =============================================
data = [
    # --- 3.3.1 Login & Logout ---
    ("3.3.1", "Pengujian Login dan Logout", "Login dengan email dan password yang valid\n(admin@siliwangi.com / password)",
     "Sistem menerima kredensial, redirect ke halaman dashboard admin, sesi pengguna aktif.",
     "Login berhasil. Admin diarahkan ke /admin/1 dan dashboard Siliwangi Admin tampil.", "✅ Lulus", "Valid"),

    ("", "", "Login dengan email tidak terdaftar\n(salah@email.com / password)",
     "Sistem menampilkan pesan validasi 'These credentials do not match our records.' dan tidak memberikan akses.",
     "Pesan error validasi muncul. Akses ditolak. Form tidak direset.", "✅ Lulus", "Valid"),

    ("", "", "Login dengan password salah\n(admin@siliwangi.com / wrongpass)",
     "Sistem menampilkan pesan error kredensial tidak cocok dan halaman login tetap tampil.",
     "Pesan error 'These credentials do not match our records.' tampil. Akses ditolak.", "✅ Lulus", "Valid"),

    ("", "", "Login dengan kolom email kosong",
     "Sistem menampilkan validasi 'The email field is required.' sebelum request dikirim.",
     "Validasi HTML5 dan Laravel muncul. Form tidak terkirim.", "✅ Lulus", "Valid"),

    ("", "", "Login dengan format email tidak valid\n(contoh: admingmail.com)",
     "Sistem menampilkan validasi 'The email field must be a valid email address.'",
     "Browser dan Laravel memvalidasi format email. Akses ditolak.", "✅ Lulus", "Valid"),

    ("", "", "Logout dari sesi aktif admin",
     "Sesi pengguna dihapus, pengguna diarahkan ke halaman login, dan tidak dapat mengakses halaman protected.",
     "Logout berhasil. Sesi dihapus. Akses ke /admin/1 diredirect ke halaman login.", "✅ Lulus", "Valid"),

    ("", "", "Login sebagai Customer\n(budi@gmail.com / password)",
     "Pengguna dengan role 'customer' berhasil login dan diarahkan ke halaman Customer Dashboard.",
     "Login berhasil. Redirect ke /dashboard dengan tampilan customer dashboard.", "✅ Lulus", "Valid"),

    # --- 3.3.2 Booking Tanpa Login ---
    ("3.3.2", "Pengujian Booking Mobil Tanpa Login", "Akses halaman katalog kendaraan tanpa login\n(/cars)",
     "Halaman katalog ditampilkan dengan semua kendaraan yang tersedia tanpa memerlukan autentikasi.",
     "Halaman /cars berhasil dimuat. Daftar 13 kendaraan ditampilkan dengan filter dan harga.", "✅ Lulus", "Valid"),

    ("", "", "Klik tombol 'Pesan Sekarang' pada kendaraan tanpa login",
     "Pengguna diarahkan ke halaman checkout dengan parameter kendaraan yang dipilih.",
     "Redirect ke /checkout/{slug} berhasil. Halaman checkout step 1 tampil.", "✅ Lulus", "Valid"),

    ("", "", "Mengisi form Step 1 (Konfigurasi Sewa) tanpa login:\nTanggal pickup, return, cabang, tipe sewa",
     "Sistem menerima input, memvalidasi ketersediaan kendaraan, dan melanjutkan ke Step 2.",
     "Step 1 berhasil divalidasi. Lanjut ke step 2 (Identitas Penyewa).", "✅ Lulus", "Valid"),

    ("", "", "Mengisi tanggal return lebih awal dari tanggal pickup",
     "Sistem menampilkan error validasi: 'The return date field must be a date after pickup date.'",
     "Validasi Laravel muncul. Step tidak dilanjutkan.", "✅ Lulus", "Valid"),

    ("", "", "Submit booking tanpa memilih cabang (branch_id kosong)",
     "Sistem menampilkan validasi 'The branch field is required.' dan tidak melanjutkan proses.",
     "Validasi muncul. Dropdown branch wajib diisi sebelum lanjut.", "✅ Lulus", "Valid"),

    ("", "", "Mengisi Step 2 (Identitas) tanpa login: nama, email, NIK, nomor SIM",
     "Sistem menerima data identitas pelanggan baru dan melanjutkan ke step 3 upload dokumen.",
     "Step 2 divalidasi. Lanjut ke step 3 upload KTP, SIM, KK.", "✅ Lulus", "Valid"),

    # --- 3.3.3 Booking Dengan Login ---
    ("3.3.3", "Pengujian Booking Mobil Dengan Login", "Login sebagai customer kemudian akses checkout kendaraan",
     "Data pengguna (nama, email, no telepon) otomatis terisi di form Step 2 dari data profil.",
     "Pre-fill data berhasil. Nama, email, dan phone customer otomatis terisi dari database.", "✅ Lulus", "Valid"),

    ("", "", "Booking kendaraan yang sudah tidak tersedia (is_available = false)",
     "Sistem menampilkan pesan error bahwa kendaraan tidak tersedia untuk tanggal yang dipilih.",
     "Pesan 'vehicle is already booked for your selected dates' muncul. Proses dihentikan.", "✅ Lulus", "Valid"),

    ("", "", "Menggunakan kode promo valid (WELCOME20) saat konfirmasi booking",
     "Sistem mengaplikasikan diskon 20% dari subtotal dan menampilkan grand total yang telah dikurangi.",
     "Kode promo WELCOME20 berhasil diterapkan. Diskon 20% terhitung pada grand total.", "✅ Lulus", "Valid"),

    ("", "", "Menggunakan kode promo yang sudah kedaluwarsa atau tidak valid",
     "Sistem menampilkan pesan error 'Kode promo tidak valid atau sudah tidak aktif.'",
     "Pesan error promo tidak valid muncul. Grand total tidak berubah.", "✅ Lulus", "Valid"),

    ("", "", "Booking dengan opsi 'Dengan Driver' dipilih",
     "Sistem menambahkan biaya driver (Rp 150.000/hari) ke kalkulasi grand total.",
     "Driver fee terhitung otomatis. Grand total bertambah sesuai jumlah hari sewa.", "✅ Lulus", "Valid"),

    ("", "", "Booking sukses – sistem membuat record booking dan payment",
     "Booking tersimpan di database dengan status 'pending', payment dibuat dengan status 'unpaid', notifikasi dikirim.",
     "Record booking dan payment berhasil dibuat. Booking code di-generate (INV-YYYYMMDD-XXXXX).", "✅ Lulus", "Valid"),

    # --- 3.3.4 Pengujian Pembayaran ---
    ("3.3.4", "Pengujian Pembayaran", "Akses halaman pembayaran Midtrans dari booking yang sudah dibuat",
     "Halaman payment gateway Midtrans terbuka dengan nominal sesuai grand total booking.",
     "Snap Midtrans berhasil dimuat dengan order_id dan nominal yang tepat.", "✅ Lulus", "Valid"),

    ("", "", "Simulasi callback Midtrans dengan status 'settlement' (pembayaran sukses)",
     "Status booking berubah menjadi 'confirmed', payment_status menjadi 'paid', payment record status 'success'.",
     "Callback diproses. Booking: confirmed, Payment: paid, Record: success. (Hasil E2E Test: PASSED)", "✅ Lulus", "Valid"),

    ("", "", "Simulasi callback Midtrans dengan status 'expire' (pembayaran kedaluwarsa)",
     "Status booking berubah menjadi 'cancelled', payment_status menjadi 'expired'.",
     "Callback expire diproses dengan benar. Status booking dan payment diperbarui sesuai.", "✅ Lulus", "Valid"),

    ("", "", "Simulasi callback Midtrans dengan signature_key tidak valid",
     "Sistem menolak callback dan tidak memperbarui status booking (keamanan notifikasi).",
     "Signature validation gagal. Callback ditolak. Status booking tidak berubah.", "✅ Lulus", "Valid"),

    ("", "", "Akses halaman invoice setelah booking dikonfirmasi (/invoice/{code})",
     "Halaman invoice menampilkan detail lengkap booking: nama penyewa, kendaraan, tanggal, dan total pembayaran.",
     "Halaman invoice berhasil dimuat dengan data booking yang sesuai.", "✅ Lulus", "Valid"),

    # --- 3.3.5 Kelola Kendaraan ---
    ("3.3.5", "Pengujian Kelola Kendaraan", "Admin mengakses halaman manajemen kendaraan di panel Filament",
     "Daftar seluruh kendaraan ditampilkan dalam tabel dengan fitur filter, search, dan sort.",
     "Halaman Cars Resource di Filament berhasil dimuat (HTTP 200). Semua kendaraan tampil.", "✅ Lulus", "Valid"),

    ("", "", "Admin menambah data kendaraan baru melalui form Create",
     "Kendaraan baru tersimpan di database, muncul di daftar, dan tersedia untuk booking.",
     "Form create kendaraan berhasil. Validasi kolom wajib (nama, plat, harga) berjalan.", "✅ Lulus", "Valid"),

    ("", "", "Admin mengubah status kendaraan menjadi 'maintenance'",
     "Status kendaraan berubah dan kendaraan tidak tersedia untuk pemesanan baru.",
     "Status berhasil diperbarui. Kendaraan tidak muncul pada daftar 'available' di katalog.", "✅ Lulus", "Valid"),

    ("", "", "Admin menghapus data kendaraan yang sedang dalam booking aktif",
     "Sistem menampilkan pesan error atau mencegah penghapusan karena ada relasi dengan booking aktif.",
     "Foreign key constraint mencegah penghapusan. Pesan error ditampilkan oleh Filament.", "✅ Lulus", "Valid"),

    ("", "", "Admin melakukan filter kendaraan berdasarkan tipe (MPV, SUV, Luxury)",
     "Hanya kendaraan dengan tipe yang dipilih yang ditampilkan dalam tabel.",
     "Filter tipe berfungsi. Hasil filter sesuai tipe kendaraan yang dipilih.", "✅ Lulus", "Valid"),

    # --- 3.3.6 Laporan Rental ---
    ("3.3.6", "Pengujian Laporan Rental", "Owner/Finance mengakses halaman laporan booking (/dashboard/reports/bookings)",
     "Laporan booking ditampilkan dengan daftar transaksi beserta detail: kendaraan, penyewa, tanggal, dan status.",
     "Halaman laporan booking berhasil dimuat. Data booking tampil dengan filter periode.", "✅ Lulus", "Valid"),

    ("", "", "Mengakses laporan pendapatan (/dashboard/reports/revenue)",
     "Laporan pendapatan menampilkan total pemasukan, breakdown per kendaraan, dan grafik tren.",
     "Halaman revenue report berhasil dimuat. Kalkulasi pendapatan sesuai data payment.", "✅ Lulus", "Valid"),

    ("", "", "Mengakses halaman analytics (/dashboard/reports/analytics)",
     "Dashboard analytics menampilkan chart statistik: bookings/bulan, kendaraan populer, occupancy rate.",
     "Analytics dashboard berhasil dimuat dengan widget Chart.js.", "✅ Lulus", "Valid"),

    ("", "", "Akses laporan oleh user dengan role 'customer' (unauthorized)",
     "Sistem menolak akses dan mengarahkan ke halaman 403 Forbidden atau redirect ke halaman yang sesuai.",
     "Middleware role:owner|finance|super-admin aktif. Akses ditolak untuk role customer.", "✅ Lulus", "Valid"),

    ("", "", "Filter laporan berdasarkan rentang tanggal tertentu",
     "Laporan hanya menampilkan data dalam rentang tanggal yang dipilih.",
     "Filter tanggal berfungsi. Data yang ditampilkan sesuai rentang yang dipilih.", "✅ Lulus", "Valid"),
]

# =============================================
# WRITE DATA
# =============================================
ROW = 5
section_fills = {
    "3.3.1": "D6E4F0",
    "3.3.2": "D9EAD3",
    "3.3.3": "FFF2CC",
    "3.3.4": "FCE5CD",
    "3.3.5": "EAD1DC",
    "3.3.6": "E8D5FF",
}
current_section = ""
no_counter = 0

for row_data in data:
    no_counter += 1
    no, skenario, tc, expected, result, status, conclusion = row_data

    if no and no != current_section:
        current_section = no

    bg = section_fills.get(current_section, "FFFFFF")

    # No
    nc = ws.cell(ROW, 1, no_counter)
    nc.font = Font(name="Calibri", bold=True, size=10)
    nc.alignment = Alignment(horizontal="center", vertical="center")
    nc.fill = PatternFill("solid", fgColor=bg)

    # Skenario (merge when first in group)
    cell_style(ws, ROW, 2, skenario, bold=(no != ""), halign="left", bg=bg)
    cell_style(ws, ROW, 3, tc, halign="left", bg=bg)
    cell_style(ws, ROW, 4, expected, halign="left", bg=bg)
    cell_style(ws, ROW, 5, result, halign="left", bg=bg)

    # Status
    sc = ws.cell(ROW, 6, status)
    sc.font = Font(name="Calibri", bold=True, size=10,
                   color="00AA44" if "Lulus" in status else "CC0000")
    sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sc.fill = PatternFill("solid", fgColor=bg)

    # Kesimpulan
    cell_style(ws, ROW, 7, conclusion, bold=False, halign="center", bg=bg)

    apply_border(ws, ROW, ROW, 1, 7)
    ws.row_dimensions[ROW].height = 60
    ROW += 1

# =============================================
# SUMMARY ROW
# =============================================
ws.row_dimensions[ROW].height = 8
ROW += 1

ws.merge_cells(f"A{ROW}:E{ROW}")
sv = ws.cell(ROW, 1, f"TOTAL PENGUJIAN: {no_counter} Test Case")
sv.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
sv.fill = PatternFill("solid", fgColor="1F3864")
sv.alignment = Alignment(horizontal="center", vertical="center")

ws.cell(ROW, 6, f"✅ {no_counter} Lulus").font = Font(bold=True, size=11, color="00AA44")
ws.cell(ROW, 6).fill = PatternFill("solid", fgColor="D9F0D9")
ws.cell(ROW, 6).alignment = Alignment(horizontal="center", vertical="center")

ws.cell(ROW, 7, "VALID").font = Font(bold=True, size=11, color="1F3864")
ws.cell(ROW, 7).fill = PatternFill("solid", fgColor="D6E4F0")
ws.cell(ROW, 7).alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[ROW].height = 28
apply_border(ws, ROW, ROW, 1, 7)

# =============================================
# FREEZE & SAVE
# =============================================
ws.freeze_panes = "A5"

out = r"documents\Pengujian_BlackBox_Siliwangi_Rental.xlsx"
wb.save(out)
print(f"[OK] File berhasil dibuat: {out}")
